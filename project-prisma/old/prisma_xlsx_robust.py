"""
PRISMA - XLSX Parser (robust)
=============================

Diese Version ist robuster gegenüber unterschiedlichen XLSX-Varianten:
- erkennt die Kopfzeile automatisch
- erkennt die Spalten dynamisch
- verarbeitet nur aktive Einträge
- blendet "Abgesetzt" konsequent aus
- versucht auch ohne Excel-Einrückung noch sinnvolle Typen zu finden

Gedanke:
Nicht jede XLSX-Datei ist ein "echter Export".
Manche entstehen durch Copy/Paste aus dem KIS in Excel.
Darauf ist diese Version besser vorbereitet.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ============================================================
# HILFSFUNKTIONEN
# ============================================================

def clean_text(value: Any) -> str:
    """Normalisiert Zellinhalte zu sauberem Text."""
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def strip_pmd_marker(title: str) -> str:
    """Entfernt die Anzeige-Markierung '(PMD)' aus einem Paketnamen."""
    return re.sub(r"\s*\(PMD\)\s*$", "", title).strip()


def parse_zyklentext(zyklentext: str) -> Dict[str, Any]:
    """Interpretiert häufige Zyklustexte in eine einfache Struktur."""
    text = clean_text(zyklentext)

    result: Dict[str, Any] = {
        "raw": text,
        "kind": None,
        "frequency_per_day": None,
        "times": [],
        "interval_hours": None,
        "notes": None,
    }

    if not text:
        return result

    lower = text.lower()

    if "bei bedarf" in lower:
        result["kind"] = "bei_bedarf"
        return result

    if lower.startswith("einmalig"):
        result["kind"] = "einmalig"
        result["notes"] = text
        return result

    match = re.search(r"(\d+)x\s*(?:tgl\.|pro tag)", lower)
    if match:
        result["kind"] = "mehrfach_pro_tag"
        result["frequency_per_day"] = int(match.group(1))
        result["times"] = re.findall(r"\b\d{1,2}:\d{2}\b", text)
        if not result["times"]:
            result["notes"] = text
        return result

    match = re.search(r"alle\s+(\d+)\s+stunden", lower)
    if match:
        result["kind"] = "stundenintervall"
        result["interval_hours"] = int(match.group(1))
        result["notes"] = text
        return result

    if lower.startswith("zyklus für die eqs-erfassung"):
        result["kind"] = "sonderzyklus"
        result["notes"] = text
        return result

    result["kind"] = "frei_text"
    result["notes"] = text
    return result


# ============================================================
# HEADER- UND SPALTENERKENNUNG
# ============================================================

HEADER_ALIASES = {
    "title": [
        "bezeichnung pflegeplaneintrag",
        "bezeichnung",
        "pflegeplaneintrag",
    ],
    "status": [
        "status",
    ],
    "evaluierung": [
        "evaluierung",
    ],
    "zyklus": [
        "zyklentext",
        "zyklus",
    ],
    "anlege_dat": [
        "anlegedat.",
        "anlegedat",
    ],
    "erf_zeit": [
        "erf.zeit",
        "erfzeit",
    ],
    "angelegt_von": [
        "angelegt von",
    ],
}


def _matches_alias(text: str, aliases: List[str]) -> bool:
    lowered = clean_text(text).lower()
    return any(alias == lowered for alias in aliases)


def detect_header_row_and_columns(ws, scan_rows: int = 25) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Sucht in den ersten Zeilen nach einer echten Kopfzeile und liefert:
    - die Zeilennummer der Kopfzeile
    - ein Mapping der erkannten Spalten

    Beispiel-Rückgabe:
        (5, {"title": 1, "status": 2, "zyklus": 4, ...})
    """
    best_row = None
    best_mapping: Dict[str, int] = {}
    best_score = -1

    max_scan_row = min(scan_rows, ws.max_row)

    for row_idx in range(1, max_scan_row + 1):
        mapping: Dict[str, int] = {}

        for col_idx in range(1, ws.max_column + 1):
            cell_text = clean_text(ws.cell(row_idx, col_idx).value)
            if not cell_text:
                continue

            for field_name, aliases in HEADER_ALIASES.items():
                if field_name in mapping:
                    continue
                if _matches_alias(cell_text, aliases):
                    mapping[field_name] = col_idx

        score = len(mapping)

        # Eine brauchbare Kopfzeile braucht mindestens Bezeichnung + Status
        if score > best_score and "title" in mapping and "status" in mapping:
            best_score = score
            best_row = row_idx
            best_mapping = mapping

    return best_row, best_mapping


# ============================================================
# TYPERKENNUNG
# ============================================================

def get_indent_level(ws, row_idx: int, title_col: int) -> int:
    """Liest die visuelle Einrückung aus der Titelspalte."""
    indent = ws.cell(row_idx, title_col).alignment.indent
    if indent is None:
        return 0
    try:
        return int(indent)
    except Exception:
        return 0


def classify_row_type_from_indent(indent_level: int) -> str:
    """
    Ordnet eine Zeile anhand der Einrückung einem Typ zu.
    Bekannte Muster:
    - 2 = Paket
    - 4 = Leistung
    - 6 = Detail
    """
    if indent_level <= 2:
        return "paket"
    if indent_level == 4:
        return "leistung"
    if indent_level >= 6:
        return "detail"
    return "unbekannt"


def infer_row_type_without_indent(title: str, prev_type: Optional[str], prev_package_title: Optional[str]) -> str:
    """
    Fallback, wenn Copy/Paste keine Excel-Einrückung mitgebracht hat.

    Heuristik:
    - Zeilen mit "(PMD)" oder führender Nummer wie "01 Bewegung" -> Paket
    - sehr kurze Einträge nach einer Leistung -> eher Detail
    - sonst standardmäßig Leistung
    """
    text = clean_text(title)

    if not text:
        return "unbekannt"

    if "(PMD)" in text:
        return "paket"

    if re.match(r"^\d{1,2}\s+\S+", text):
        return "paket"

    if re.match(r"^\d{1,2}[.)-]\s+\S+", text):
        return "paket"

    # Typische Parameter-/Detailmuster
    if len(text) <= 35 and "[" in text and "]" in text:
        return "detail"

    if prev_type == "leistung" and len(text) <= 40:
        return "detail"

    if prev_type == "paket" and prev_package_title and text != prev_package_title:
        return "leistung"

    return "leistung"


# ============================================================
# PARSER
# ============================================================

def parse_pflegeplan_xlsx(xlsx_path: str | Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Liest den Pflegeplan aus einer Excel-Datei und gibt zurück:
    1. aktive Einträge
    2. Parser-Statistik
    """
    workbook = load_workbook(xlsx_path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row, col_map = detect_header_row_and_columns(worksheet)

    # Fallback: Wenn keine Kopfzeile gefunden wurde, nehmen wir die klassische Struktur an
    if header_row is None:
        header_row = 5
        col_map = {
            "title": 1,
            "status": 2,
            "evaluierung": 3,
            "zyklus": 4,
            "anlege_dat": 5,
            "erf_zeit": 6,
            "angelegt_von": 7,
        }

    title_col = col_map.get("title", 1)
    status_col = col_map.get("status", 2)
    zyklus_col = col_map.get("zyklus", 4)

    first_data_row = header_row + 1

    parsed_rows: List[Dict[str, Any]] = []

    current_package: Optional[Dict[str, Any]] = None
    current_leistung: Optional[Dict[str, Any]] = None
    prev_type: Optional[str] = None

    stats: Dict[str, Any] = {
        "sheet_name": worksheet.title,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "header_row_detected": header_row,
        "first_data_row": first_data_row,
        "columns_detected": col_map,
        "rows_total_scanned": 0,
        "rows_empty": 0,
        "rows_ignored_no_title": 0,
        "rows_abgesetzt": 0,
        "rows_aktiv": 0,
        "rows_other_status": 0,
        "entries_output": 0,
    }

    for row_idx in range(first_data_row, worksheet.max_row + 1):
        raw_title = worksheet.cell(row_idx, title_col).value
        raw_status = worksheet.cell(row_idx, status_col).value
        raw_zyklus = worksheet.cell(row_idx, zyklus_col).value if zyklus_col <= worksheet.max_column else None

        title = clean_text(raw_title)
        status = clean_text(raw_status)
        status_lower = status.lower()

        if not title and not status:
            stats["rows_empty"] += 1
            continue

        stats["rows_total_scanned"] += 1

        # Metazeilen oder erneut eingefügte Header überspringen
        if "bezeichnung pflegeplaneintrag" in title.lower():
            continue

        if not title:
            stats["rows_ignored_no_title"] += 1
            continue

        # ----------------------------------------------------
        # STATUSFILTER
        # ----------------------------------------------------
        if "abgesetzt" in status_lower:
            stats["rows_abgesetzt"] += 1
            continue

        if "aktiv" not in status_lower:
            stats["rows_other_status"] += 1
            continue

        stats["rows_aktiv"] += 1

        # ----------------------------------------------------
        # ZEILENKLASSIFIKATION
        # ----------------------------------------------------
        indent_level = get_indent_level(worksheet, row_idx, title_col)

        if indent_level > 0:
            row_type = classify_row_type_from_indent(indent_level)
        else:
            row_type = infer_row_type_without_indent(
                title=title,
                prev_type=prev_type,
                prev_package_title=current_package["title_clean"] if current_package else None,
            )

        zyklus_raw = clean_text(raw_zyklus)
        zyklus_info = parse_zyklentext(zyklus_raw)

        # ----------------------------------------------------
        # KONTEXT AKTUALISIEREN
        # ----------------------------------------------------
        if row_type == "paket":
            is_pmd = "(PMD)" in title
            current_package = {
                "title_raw": title,
                "title_clean": strip_pmd_marker(title),
                "is_pmd": is_pmd,
                "row": row_idx,
            }
            current_leistung = None

        elif row_type == "leistung":
            current_leistung = {
                "title": title,
                "row": row_idx,
            }

        entry = {
            "row_index": row_idx,
            "title": title,
            "status": status,
            "row_type": row_type,
            "indent_level": indent_level,
            "package": current_package["title_clean"] if current_package else None,
            "package_raw": current_package["title_raw"] if current_package else None,
            "package_is_pmd": current_package["is_pmd"] if current_package else False,
            "entry_origin": "PMD" if (current_package and current_package["is_pmd"]) else "manuell",
            "parent_leistung": current_leistung["title"] if (row_type == "detail" and current_leistung) else None,
            "zyklentext": zyklus_raw,
            "zyklus_info": zyklus_info,
            "is_planned_core": row_type in {"paket", "leistung"},
        }

        parsed_rows.append(entry)
        prev_type = row_type

    stats["entries_output"] = len(parsed_rows)
    return parsed_rows, stats


def export_to_json(data: List[Dict[str, Any]], output_path: str | Path) -> None:
    """Speichert die Parser-Ausgabe als JSON-Datei."""
    Path(output_path).write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


if __name__ == "__main__":
    source_file = Path("TEst Pflegeplan.xlsx")
    output_file = Path("pflegeplan_parsed_preview.json")

    entries, stats = parse_pflegeplan_xlsx(source_file)
    export_to_json(entries, output_file)

    print(f"Fertig. {len(entries)} aktive Einträge exportiert nach: {output_file}")
    print("Parser-Statistik:")
    for key, value in stats.items():
        print(f"  {key}: {value}")
