"""
PRISMA - Parserkern
===================

Diese Datei enthält die Kernlogik zum Einlesen und Vorverarbeiten
eines Pflegeplans aus zwei Quellen:

1. Excel-Export (.xlsx)
2. direkt eingefügter KIS-Text (tabellarischer Copy/Paste)

Wichtige Regeln:
- Status "Abgesetzt" wird ignoriert
- "Angelegt von" wird ignoriert
- Zyklentext wird berücksichtigt
- (PMD) wird auf Paketebene erkannt
- Einrückungsebene 6 wird NICHT als eigene geplante Leistung gezählt,
  sondern als Parameter/Zusatz einer Leistung behandelt
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


# ============================================================
# KONFIGURATION
# ============================================================

HEADER_ROW = 5
FIRST_DATA_ROW = 6

COL_TITLE = 1
COL_STATUS = 2
COL_EVAL = 3
COL_ZYKLUS = 4
COL_ANLEGEDAT = 5
COL_ERFZEIT = 6
COL_ANGELEGT_VON = 7


# ============================================================
# HILFSFUNKTIONEN - TEXT
# ============================================================

def clean_text(value: Any) -> str:
    """Normalisiert beliebige Zell-/Texteingaben zu einem sauberen String."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\u2007", " ").replace("\u202f", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def strip_pmd_marker(title: str) -> str:
    """Entfernt '(PMD)' nur aus der Anzeige, nicht aus der Logik."""
    return re.sub(r"\s*\(PMD\)\s*$", "", title).strip()


def get_leading_indent_hint(raw_title: Any) -> int:
    """
    Ermittelt eine grobe Einrückung aus kopiertem Rohtext.

    Beim .xlsx-Export steckt die Struktur in der Zellformatierung.
    Beim Copy/Paste aus dem KIS können manchmal echte Leerzeichen
    oder Tabs am Anfang landen. Diese Funktion versucht daraus
    einen Hinweis zu gewinnen.
    """
    if raw_title is None:
        return 0

    text = str(raw_title).replace("\xa0", " ")
    match = re.match(r"^([ \t]+)", text)
    if not match:
        return 0

    prefix = match.group(1)
    spaces = prefix.count(" ") + prefix.count("\t") * 4

    if spaces >= 6:
        return 6
    if spaces >= 4:
        return 4
    if spaces >= 2:
        return 2
    return 0


# ============================================================
# HILFSFUNKTIONEN - STRUKTUR
# ============================================================

def classify_row_type(indent_level: int) -> str:
    """
    Ordnet die Struktur einer Zeile einer fachlichen Art zu.

    Wichtige Änderung:
    - Ebene 6+ wird als 'parameter' statt als 'detail' behandelt,
      weil diese Zeilen im Export oft nur Zusätze/Unterparameter einer
      eigentlichen Leistung sind, nicht eigenständig geplante Leistungen.
    """
    if indent_level <= 2:
        return "paket"
    if indent_level == 4:
        return "leistung"
    if indent_level >= 6:
        return "parameter"
    return "unbekannt"


def infer_row_type_without_indent(
    title: str,
    zyklentext: str,
) -> str:
    """
    Fallback-Heuristik für Copy/Paste-Text, wenn keine Einrückung mitgeliefert wird.

    Prinzip:
    - Paket nur dann, wenn wir relativ sicher sind
    - sonst lieber als Leistung behandeln, statt künstlich Details zu erfinden
    """
    title_clean = clean_text(title)
    zyklus_clean = clean_text(zyklentext)

    if "(PMD)" in title_clean:
        return "paket"

    if re.match(r"^\d{2}\s+", title_clean):
        return "paket"

    if (
        not zyklus_clean
        and len(title_clean.split()) <= 4
        and not re.search(r"\b(durchführen|verabreichen|überwachen|messen|wechseln|reinigen|bereitstellen)\b", title_clean, re.I)
        and "[" not in title_clean
        and ":" not in title_clean
    ):
        return "paket"

    return "leistung"


# ============================================================
# HILFSFUNKTIONEN - ZYKLUS
# ============================================================

def parse_zyklentext(zyklentext: str) -> Dict[str, Any]:
    """Erkennt häufige Zyklusmuster und gibt sie strukturiert zurück."""
    text = clean_text(zyklentext)

    result: Dict[str, Any] = {
        "raw": text,
        "kind": None,
        "frequency_per_day": None,
        "times": [],
        "interval_hours": None,
        "notes": None,
    }

    if not text:
        return result

    lower = text.lower()

    if "bei bedarf" in lower:
        result["kind"] = "bei_bedarf"
        return result

    if lower.startswith("einmalig"):
        result["kind"] = "einmalig"
        result["notes"] = text
        return result

    match = re.search(r"(\d+)x\s*(?:tgl\.|pro tag)", lower)
    if match:
        result["kind"] = "mehrfach_pro_tag"
        result["frequency_per_day"] = int(match.group(1))
        result["times"] = re.findall(r"\b\d{1,2}:\d{2}\b", text)
        if not result["times"]:
            result["notes"] = text
        return result

    match = re.search(r"alle\s+(\d+)\s+stunden", lower)
    if match:
        result["kind"] = "stundenintervall"
        result["interval_hours"] = int(match.group(1))
        result["notes"] = text
        return result

    if lower.startswith("zyklus für die eqs-erfassung"):
        result["kind"] = "sonderzyklus"
        result["notes"] = text
        return result

    result["kind"] = "frei_text"
    result["notes"] = text
    return result


# ============================================================
# NORMALISIERTE ZEILEN IN EINTRÄGE UMWANDELN
# ============================================================

def _build_entries_from_normalized_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Vereinheitlicht die Verarbeitung für Excel und Copy/Paste."""
    parsed_rows: List[Dict[str, Any]] = []

    current_package: Optional[Dict[str, Any]] = None
    current_leistung: Optional[Dict[str, Any]] = None

    for row in rows:
        title_raw = clean_text(row.get("title_raw"))
        status = clean_text(row.get("status"))

        if not title_raw:
            continue

        if "abgesetzt" in status.lower():
            continue

        indent_level = int(row.get("indent_level", 0) or 0)
        explicit_row_type = row.get("row_type")

        zyklus_raw = clean_text(row.get("zyklentext"))
        zyklus_info = parse_zyklentext(zyklus_raw)

        if explicit_row_type:
            row_type = explicit_row_type
        else:
            if indent_level > 0:
                row_type = classify_row_type(indent_level)
            else:
                row_type = infer_row_type_without_indent(
                    title=title_raw,
                    zyklentext=zyklus_raw,
                )

        if row_type == "paket":
            is_pmd = "(PMD)" in title_raw
            current_package = {
                "title_raw": title_raw,
                "title_clean": strip_pmd_marker(title_raw),
                "is_pmd": is_pmd,
                "row": row.get("row_index"),
            }
            current_leistung = None

        elif row_type == "leistung":
            current_leistung = {
                "title": title_raw,
                "row": row.get("row_index"),
            }

        entry = {
            "row_index": row.get("row_index"),
            "title": title_raw,
            "status": status,
            "row_type": row_type,
            "indent_level": indent_level,
            "package": current_package["title_clean"] if current_package else None,
            "package_raw": current_package["title_raw"] if current_package else None,
            "package_is_pmd": current_package["is_pmd"] if current_package else False,
            "entry_origin": "PMD" if (current_package and current_package["is_pmd"]) else "manuell",
            "parent_leistung": current_leistung["title"] if (row_type == "parameter" and current_leistung) else None,
            "zyklentext": zyklus_raw,
            "zyklus_info": zyklus_info,
            "source_mode": row.get("source_mode", "unknown"),
            "is_planned_item": row_type in {"paket", "leistung"},
        }

        parsed_rows.append(entry)

    return parsed_rows


# ============================================================
# EXCEL-PARSER
# ============================================================

def parse_pflegeplan_xlsx(xlsx_path: str | Path) -> List[Dict[str, Any]]:
    """Liest einen Pflegeplan aus einer Excel-Datei ein."""
    workbook = load_workbook(xlsx_path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]

    normalized_rows: List[Dict[str, Any]] = []

    for row_idx in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        raw_title = worksheet.cell(row_idx, COL_TITLE).value
        status = clean_text(worksheet.cell(row_idx, COL_STATUS).value)

        if raw_title is None and not status:
            continue

        if not raw_title:
            continue

        indent = worksheet.cell(row_idx, COL_TITLE).alignment.indent
        indent_level = int(indent) if indent is not None else 0

        normalized_rows.append(
            {
                "row_index": row_idx,
                "title_raw": raw_title,
                "status": worksheet.cell(row_idx, COL_STATUS).value,
                "evaluierung": worksheet.cell(row_idx, COL_EVAL).value,
                "zyklentext": worksheet.cell(row_idx, COL_ZYKLUS).value,
                "indent_level": indent_level,
                "row_type": classify_row_type(indent_level),
                "source_mode": "xlsx",
            }
        )

    return _build_entries_from_normalized_rows(normalized_rows)


# ============================================================
# TEXT-PARSER
# ============================================================

def _split_pasted_line(line: str) -> List[str]:
    """
    Zerlegt eine eingefügte Zeile in möglichst sinnvolle Spalten.

    Priorität:
    1. Tabs
    2. viele Leerzeichen als Spaltentrenner
    3. sonst Rückgabe als Einzelstring

    Wichtig:
    Wir filtern leere Splitteile nicht aggressiv weg, damit
    Spaltenpositionen möglichst erhalten bleiben.
    """
    if "\t" in line:
        return [part.strip() for part in line.split("\t")]

    parts = re.split(r"\s{2,}", line.rstrip())

    if len(parts) > 1:
        return [part.strip() for part in parts]

    return [line.rstrip()]


def _extract_status_from_line(parts: List[str], full_line: str) -> str:
    """
    Versucht den Status robust aus einer eingefügten Textzeile zu erkennen.

    Warum das nötig ist:
    Im KIS-Copy/Paste landet der Status nicht immer sauber in Spalte 2.
    Deshalb prüfen wir sowohl die erkannten Teile als auch die gesamte Zeile.
    """
    candidates = [clean_text(p) for p in parts if clean_text(p)]

    for candidate in candidates:
        lower = candidate.lower()
        if "abgesetzt" in lower:
            return "Abgesetzt"
        if lower == "aktiv":
            return "Aktiv"

    full_lower = clean_text(full_line).lower()

    if "abgesetzt" in full_lower:
        return "Abgesetzt"
    if re.search(r"\baktiv\b", full_lower):
        return "Aktiv"

    return ""


def parse_pflegeplan_text(raw_text: str) -> List[Dict[str, Any]]:
    """
    Liest direkt eingefügten Text aus dem KIS ein.

    Diese Version ist robuster gegen unsaubere Copy/Paste-Strukturen:
    - erkennt 'Abgesetzt' auch dann, wenn der Status nicht sauber in einer Spalte landet
    - ignoriert 'Angelegt von'
    - ignoriert offensichtliche Kopf-/Meta-Zeilen
    """
    lines = [line.rstrip("\n\r") for line in str(raw_text or "").splitlines()]
    lines = [line for line in lines if clean_text(line)]

    normalized_rows: List[Dict[str, Any]] = []
    row_counter = 0

    for line in lines:
        plain = clean_text(line)
        plain_lower = plain.lower()

        # ----------------------------------------------------
        # META- UND KOPFZEILEN IGNORIEREN
        # ----------------------------------------------------
        if plain.startswith("Pfleg OE:"):
            continue

        if "Fachabt.:" in plain:
            continue

        if "Bezeichnung Pflegeplaneintrag" in plain and "Status" in plain:
            continue

        if plain_lower.startswith("angelegt von"):
            continue

        if plain_lower.startswith("anlegedat"):
            continue

        if plain_lower.startswith("erf.zeit"):
            continue

        # ----------------------------------------------------
        # ZEILE AUFBRECHEN
        # ----------------------------------------------------
        parts = _split_pasted_line(line)

        # Falls nur eine einzige Spalte erkannt wird,
        # behandeln wir die komplette Zeile erstmal als Titel.
        title_raw = clean_text(parts[0]) if parts else ""
        status = _extract_status_from_line(parts, line)

        evaluierung = ""
        zyklentext = ""

        # Wenn Spalten sauber erkannt wurden, nehmen wir die üblichen Positionen.
        if len(parts) > 2:
            evaluierung = clean_text(parts[2])

        if len(parts) > 3:
            zyklentext = clean_text(parts[3])

        # ----------------------------------------------------
        # ZEILEN OHNE SUBSTANZ IGNORIEREN
        # ----------------------------------------------------
        if not title_raw:
            continue

        # Reine Status-/Restzeilen ohne echte Bezeichnung überspringen
        if title_raw.lower() in {"aktiv", "abgesetzt"}:
            continue

        # ----------------------------------------------------
        # EINRÜCKUNG / TYP
        # ----------------------------------------------------
        row_counter += 1
        indent_hint = get_leading_indent_hint(parts[0] if parts else line)

        normalized_rows.append(
            {
                "row_index": row_counter,
                "title_raw": title_raw,
                "status": status,
                "evaluierung": evaluierung,
                "zyklentext": zyklentext,
                "indent_level": indent_hint,
                "row_type": classify_row_type(indent_hint) if indent_hint > 0 else None,
                "source_mode": "text",
            }
        )

    return _build_entries_from_normalized_rows(normalized_rows)


# ============================================================
# EXPORT
# ============================================================

def export_to_json(data: List[Dict[str, Any]], output_path: str | Path) -> None:
    """Speichert beliebige Parser-Daten als JSON-Datei."""
    Path(output_path).write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


# ============================================================
# KLEINER TESTLAUF
# ============================================================

if __name__ == "__main__":
    source_file = Path("TEst Pflegeplan.xlsx")
    output_file = Path("pflegeplan_parsed_preview.json")

    parsed_entries = parse_pflegeplan_xlsx(source_file)
    export_to_json(parsed_entries, output_file)

    print(f"Fertig. {len(parsed_entries)} aktive Einträge exportiert nach: {output_file}")
