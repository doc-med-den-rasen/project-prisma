"""
PRISMA - XLSX Parser für Zweizeilen-Format
==========================================

Diese Version ist speziell auf die aus dem KIS heraus gespeicherte
Zweizeilen-Struktur abgestimmt:

- Zeile N   = Titel / Pflegeplaneintrag
- Zeile N+1 = Status + Zyklus + Metadaten

Wichtig:
Die Datei ist NICHT als klassisches Tabellenformat aufgebaut.
Darauf reagiert dieser Parser gezielt.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ============================================================
# HILFSFUNKTIONEN
# ============================================================

def clean_text(value: Any) -> str:
    """Normalisiert Zellinhalte zu sauberem Text."""
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def strip_pmd_marker(title: str) -> str:
    """Entfernt '(PMD)' aus der Anzeige, ohne die Information zu verlieren."""
    return re.sub(r"\s*\(PMD\)\s*$", "", title).strip()


def parse_zyklentext(zyklentext: str) -> Dict[str, Any]:
    """Interpretiert häufige Zyklustexte in eine einfache Struktur."""
    text = clean_text(zyklentext)

    result: Dict[str, Any] = {
        "raw": text,
        "kind": None,
        "frequency_per_day": None,
        "times": [],
        "interval_hours": None,
        "notes": None,
    }

    if not text:
        return result

    lower = text.lower()

    if "bei bedarf" in lower:
        result["kind"] = "bei_bedarf"
        return result

    if lower.startswith("einmalig"):
        result["kind"] = "einmalig"
        result["notes"] = text
        return result

    match = re.search(r"(\d+)x\s*(?:tgl\.|pro tag)", lower)
    if match:
        result["kind"] = "mehrfach_pro_tag"
        result["frequency_per_day"] = int(match.group(1))
        result["times"] = re.findall(r"\b\d{1,2}:\d{2}\b", text)
        if not result["times"]:
            result["notes"] = text
        return result

    match = re.search(r"alle\s+(\d+)\s+stunden", lower)
    if match:
        result["kind"] = "stundenintervall"
        result["interval_hours"] = int(match.group(1))
        result["notes"] = text
        return result

    if lower.startswith("zyklus für die eqs-erfassung"):
        result["kind"] = "sonderzyklus"
        result["notes"] = text
        return result

    result["kind"] = "frei_text"
    result["notes"] = text
    return result


# ============================================================
# HEADER-SUCHE
# ============================================================

def detect_header_row(ws, max_scan_rows: int = 30) -> Optional[int]:
    """
    Sucht die Kopfzeile, die 'Bezeichnung Pflegeplaneintrag' enthält.
    """
    upper = min(max_scan_rows, ws.max_row)

    for row_idx in range(1, upper + 1):
        first = clean_text(ws.cell(row_idx, 1).value).lower()
        if "bezeichnung pflegeplaneintrag" in first:
            return row_idx

    return None


# ============================================================
# HEURISTIKEN FÜR TYPEN
# ============================================================

_ACTION_MARKERS = [
    " durchführen",
    " unterstützen",
    " überwachen",
    " verabreichen",
    " wechseln",
    " zurechtmachen",
    " assistieren",
    " bereitstellen/abräumen",
    " führen",
    " messen",
    " leeren/wechseln",
    " reichen/entfernen",
    " anziehen",
    " ausziehen",
    " anlegen",
    " begleiten",
    " mobilisieren",
    " lagern",
    " kontrollieren",
    " versorgen",
    " vorbereiten",
    " vor-/nachbereiten",
    " punktieren",
    " absaugen",
    " planen",
]


def is_action_title(title: str) -> bool:
    """
    Erkennt typische Leistungs-/Aktionsformulierung.
    """
    t = clean_text(title).lower()

    if any(marker.strip() in t for marker in _ACTION_MARKERS):
        return True

    if t.startswith("visite "):
        return True

    if t.startswith("nächtlicher rundgang"):
        return True

    return False


def is_package_title(title: str, zyklus: str, next_title: str, prev_type: Optional[str]) -> bool:
    """
    Heuristik für Paket-/Bereichszeilen.

    Sicher erkannt:
    - nummerierte Hauptbereiche wie '10 Dekubitus/Wunde (PMD)'
    - Zeilen mit '(PMD)'

    Zusätzlich vorsichtige Heuristik für manuelle Pakete wie:
    - 'Tagesroutine'
    - 'Infusionen 15/18/25'
    - 'VAC 15/18/25'
    - 'EQS Pneu'
    """
    t = clean_text(title)
    tl = t.lower()
    next_t = clean_text(next_title)

    if not t:
        return False

    if "(pmd)" in tl:
        return True

    if re.match(r"^\d{1,2}\s+\S+", t):
        return True

    if zyklus:
        return False

    if tl.startswith("mit "):
        return False

    if "[" in t and "]" in t:
        return False

    if ":" in t and not re.match(r"^\d{1,2}\s+\S+", t):
        return False

    if is_action_title(t):
        return False

    explicit_package_keywords = [
        "routine",
        "geräte und hilfsmittel",
        "eqs ",
        "infusionen",
        "vac ",
    ]
    if any(keyword in tl for keyword in explicit_package_keywords):
        return True

    if prev_type in (None, "detail") and next_t:
        if is_action_title(next_t) or "[" in next_t or ":" in next_t or next_t.lower().startswith("mit "):
            return True

    return False


def is_detail_title(title: str, zyklus: str, prev_type: Optional[str]) -> bool:
    """
    Heuristik für Detailebene / Zusatzparameter.
    """
    t = clean_text(title)
    tl = t.lower()

    if not t:
        return False

    if tl.startswith("mit "):
        return True

    if zyklus:
        return False

    if prev_type in ("leistung", "detail"):
        if not is_action_title(t):
            return True

    return False


# ============================================================
# PARSER
# ============================================================

def parse_pflegeplan_xlsx_paired(xlsx_path: str | Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Liest eine XLSX-Datei im KIS-Zweizeilen-Format und gibt zurück:
    1. aktive, gepaarte Einträge
    2. Parser-Statistik
    """
    workbook = load_workbook(xlsx_path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = detect_header_row(worksheet)
    if header_row is None:
        raise ValueError("Kopfzeile mit 'Bezeichnung Pflegeplaneintrag' wurde nicht gefunden.")

    first_data_row = header_row + 1

    parsed_entries: List[Dict[str, Any]] = []

    current_package: Optional[Dict[str, Any]] = None
    current_leistung: Optional[Dict[str, Any]] = None
    prev_type: Optional[str] = None

    stats: Dict[str, Any] = {
        "sheet_name": worksheet.title,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "header_row_detected": header_row,
        "first_data_row": first_data_row,
        "logical_pairs_total": 0,
        "logical_pairs_aktiv": 0,
        "logical_pairs_abgesetzt": 0,
        "rows_skipped_unpaired": 0,
        "entries_output": 0,
        "format_note": "Titel steht in Zeile N, Status/Zyklus in Zeile N+1",
    }

    row_idx = first_data_row

    while row_idx <= worksheet.max_row:
        title = clean_text(worksheet.cell(row_idx, 1).value)

        if not title:
            row_idx += 1
            continue

        if "bezeichnung pflegeplaneintrag" in title.lower():
            row_idx += 1
            continue

        next_status = clean_text(worksheet.cell(row_idx + 1, 1).value) if row_idx + 1 <= worksheet.max_row else ""
        next_status_lower = next_status.lower()

        # ----------------------------------------------------
        # ECHTES ZWEIZEILEN-PAAR?
        # ----------------------------------------------------
        if next_status_lower not in {"aktiv", "abgesetzt"}:
            stats["rows_skipped_unpaired"] += 1
            row_idx += 1
            continue

        stats["logical_pairs_total"] += 1

        zyklentext = clean_text(worksheet.cell(row_idx + 1, 3).value) if row_idx + 1 <= worksheet.max_row else ""
        anlegedat = worksheet.cell(row_idx + 1, 4).value if row_idx + 1 <= worksheet.max_row else None
        erfzeit = worksheet.cell(row_idx + 1, 5).value if row_idx + 1 <= worksheet.max_row else None
        angelegt_von = clean_text(worksheet.cell(row_idx + 1, 6).value) if row_idx + 1 <= worksheet.max_row else ""

        if next_status_lower == "abgesetzt":
            stats["logical_pairs_abgesetzt"] += 1
            row_idx += 2
            continue

        stats["logical_pairs_aktiv"] += 1

        # ----------------------------------------------------
        # LOOKAHEAD für Typ-Heuristik
        # ----------------------------------------------------
        next_title_candidate = ""
        lookahead_row = row_idx + 2

        while lookahead_row <= worksheet.max_row:
            candidate_title = clean_text(worksheet.cell(lookahead_row, 1).value)
            candidate_status = clean_text(worksheet.cell(lookahead_row + 1, 1).value) if lookahead_row + 1 <= worksheet.max_row else ""

            if candidate_title and candidate_status.lower() in {"aktiv", "abgesetzt"}:
                next_title_candidate = candidate_title
                break

            lookahead_row += 1

        # ----------------------------------------------------
        # TYP HEURISTISCH KLASSIFIZIEREN
        # ----------------------------------------------------
        if is_package_title(title, zyklentext, next_title_candidate, prev_type):
            row_type = "paket"
        elif is_detail_title(title, zyklentext, prev_type):
            row_type = "detail"
        else:
            row_type = "leistung"

        zyklus_info = parse_zyklentext(zyklentext)

        if row_type == "paket":
            is_pmd = "(PMD)" in title
            current_package = {
                "title_raw": title,
                "title_clean": strip_pmd_marker(title),
                "is_pmd": is_pmd,
                "row": row_idx,
            }
            current_leistung = None

        elif row_type == "leistung":
            current_leistung = {
                "title": title,
                "row": row_idx,
            }

        entry = {
            "title_row": row_idx,
            "status_row": row_idx + 1,
            "title": title,
            "status": next_status,
            "row_type": row_type,
            "package": current_package["title_clean"] if current_package else None,
            "package_raw": current_package["title_raw"] if current_package else None,
            "package_is_pmd": current_package["is_pmd"] if current_package else False,
            "entry_origin": "PMD" if (current_package and current_package["is_pmd"]) else "manuell",
            "parent_leistung": current_leistung["title"] if (row_type == "detail" and current_leistung) else None,
            "zyklentext": zyklentext,
            "zyklus_info": zyklus_info,
            "anlegedat": anlegedat,
            "erfzeit": erfzeit,
            "angelegt_von": angelegt_von,
            "is_planned_core": row_type in {"paket", "leistung"},
        }

        parsed_entries.append(entry)
        prev_type = row_type
        row_idx += 2

    stats["entries_output"] = len(parsed_entries)
    return parsed_entries, stats


def export_to_json(data: List[Dict[str, Any]], output_path: str | Path) -> None:
    """Speichert die Parser-Ausgabe als JSON-Datei."""
    Path(output_path).write_text(
        json.dumps(data, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )


if __name__ == "__main__":
    print("Dies ist die Parser-Datei von PRISMA.")
    print("Bitte starte die Streamlit-App-Datei und nicht diese Datei direkt.")
