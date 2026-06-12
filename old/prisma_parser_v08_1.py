"""
PRISMA - Parser v08
===================

v08-Schwerpunkt:
- übernimmt die robuste Zweizeilen-Logik aus v07
- vergibt eindeutige IDs für Pakete, Leistungen und Details
- Details hängen an einer *konkreten Leistungsinstanz*, nicht nur am Titel
- dadurch können gleichnamige Leistungen im selben Paket sauber getrennt werden
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# ============================================================
# KONSTANTEN / REGELN
# ============================================================

PAIR_COL_TITLE = 1
PAIR_COL_STATUS = 1
PAIR_COL_ZYKLUS = 3
PAIR_COL_ANLEGEDAT = 4
PAIR_COL_ERFZEIT = 5
PAIR_COL_ANGELEGT_VON = 6

MANUAL_PACKAGE_EXACT = {
    "tagesroutine",
    "aufnahme",
    "op vorbereitung",
    "op-vorbereitung",
    "geräte und hilfsmittel",
    "ppr 2.0 zusatzdaten",
    "eqs pneu",
    "verlegung",
}

MANUAL_PACKAGE_PATTERNS = [
    r"^(infusionen|vac|redon|perfusoren)\s+15/18/25$",
]

ACTION_MARKERS = [
    " durchführen",
    " unterstützen",
    " überwachen",
    " verabreichen",
    " wechseln",
    " zurechtmachen",
    " assistieren",
    " bereitstellen/abräumen",
    " führen",
    " messen",
    " leeren/wechseln",
    " reichen/entfernen",
    " anziehen",
    " ausziehen",
    " anlegen",
    " begleiten",
    " mobilisieren",
    " lagern",
    " kontrollieren",
    " versorgen",
    " vorbereiten",
    " vor-/nachbereiten",
    " punktieren",
    " absaugen",
    " planen",
    " organisieren",
    " erheben",
    " abnehmen",
    " berechnen",
    " entfernen",
    " ein-/auspacken",
    " nachbereiten",
    " überprüfen/warten",
    " beobachten",
]

# ============================================================
# HILFSFUNKTIONEN
# ============================================================

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def strip_pmd_marker(title: str) -> str:
    return re.sub(r"\s*\(PMD\)\s*$", "", clean_text(title), flags=re.IGNORECASE).strip()


def normalize_title_key(title: str) -> str:
    return clean_text(title).lower()


def package_sort_key(package_name: str) -> tuple:
    t = strip_pmd_marker(package_name)
    match = re.match(r"^(\d{1,2})\s+(.+)$", t)
    if match:
        return (0, int(match.group(1)), match.group(2).lower())
    return (1, 999, t.lower())


def parse_zyklentext(zyklentext: str) -> Dict[str, Any]:
    text = clean_text(zyklentext)

    result: Dict[str, Any] = {
        "raw": text,
        "kind": None,
        "frequency_per_day": None,
        "times": [],
        "interval_hours": None,
        "notes": None,
    }

    if not text:
        return result

    lower = text.lower()

    if "bei bedarf" in lower:
        result["kind"] = "bei_bedarf"
        return result

    if lower.startswith("einmalig"):
        result["kind"] = "einmalig"
        result["notes"] = text
        return result

    match = re.search(r"(\d+)x\s*(?:tgl\.|pro tag)", lower)
    if match:
        result["kind"] = "mehrfach_pro_tag"
        result["frequency_per_day"] = int(match.group(1))
        result["times"] = re.findall(r"\b\d{1,2}:\d{2}\b", text)
        if not result["times"]:
            result["notes"] = text
        return result

    match = re.search(r"alle\s+(\d+)\s+stunden", lower)
    if match:
        result["kind"] = "stundenintervall"
        result["interval_hours"] = int(match.group(1))
        result["notes"] = text
        return result

    if lower.startswith("zyklus für die eqs-erfassung"):
        result["kind"] = "sonderzyklus"
        result["notes"] = text
        return result

    result["kind"] = "frei_text"
    result["notes"] = text
    return result


def detect_header_row(ws, scan_rows: int = 15) -> int:
    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        first_cell = clean_text(ws.cell(row_idx, 1).value).lower()
        if first_cell.startswith("bezeichnung pflegeplaneintrag"):
            return row_idx
    return 1


def is_package_title(title: str) -> bool:
    t = clean_text(title)
    tl = t.lower()

    if not t:
        return False

    if "(pmd)" in tl:
        return True

    if tl in MANUAL_PACKAGE_EXACT:
        return True

    if tl.startswith("eqs "):
        return True

    return any(re.match(pattern, tl) for pattern in MANUAL_PACKAGE_PATTERNS)


def is_action_title(title: str) -> bool:
    t = clean_text(title).lower()

    if not t:
        return False

    if any(marker.strip() in t for marker in ACTION_MARKERS):
        return True

    if t.startswith("visite "):
        return True

    if t.startswith("nächtlicher rundgang"):
        return True

    return False


def extract_logical_pairs(ws) -> tuple[list[dict[str, Any]], int]:
    header_row = detect_header_row(ws)
    first_data_row = header_row + 1

    logical_pairs: list[dict[str, Any]] = []

    row_idx = first_data_row
    while row_idx <= ws.max_row:
        title = clean_text(ws.cell(row_idx, PAIR_COL_TITLE).value)

        if not title:
            row_idx += 1
            continue

        if title.lower().startswith("bezeichnung pflegeplaneintrag"):
            row_idx += 1
            continue

        status_row = row_idx + 1
        status = clean_text(ws.cell(status_row, PAIR_COL_STATUS).value) if status_row <= ws.max_row else ""
        zyklus = clean_text(ws.cell(status_row, PAIR_COL_ZYKLUS).value) if status_row <= ws.max_row else ""
        anlegedat = clean_text(ws.cell(status_row, PAIR_COL_ANLEGEDAT).value) if status_row <= ws.max_row else ""
        erfzeit = clean_text(ws.cell(status_row, PAIR_COL_ERFZEIT).value) if status_row <= ws.max_row else ""
        angelegt_von = clean_text(ws.cell(status_row, PAIR_COL_ANGELEGT_VON).value) if status_row <= ws.max_row else ""

        logical_pairs.append(
            {
                "title_row": row_idx,
                "status_row": status_row,
                "title": title,
                "status": status,
                "zyklentext": zyklus,
                "anlegedat": anlegedat,
                "erfzeit": erfzeit,
                "angelegt_von": angelegt_von,
            }
        )

        row_idx += 2

    return logical_pairs, header_row


# ============================================================
# ID-HILFE
# ============================================================

def _next_id(counter: dict[str, int]) -> str:
    counter["value"] += 1
    return f"E{counter['value']:05d}"


# ============================================================
# KERNPARSER
# ============================================================

def parse_pflegeplan_xlsx_v08_1(xlsx_path: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(xlsx_path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]

    logical_pairs, header_row = extract_logical_pairs(worksheet)

    active_pairs = [pair for pair in logical_pairs if pair["status"].lower() == "aktiv"]
    abgesetzt_pairs = [pair for pair in logical_pairs if "abgesetzt" in pair["status"].lower()]

    package_history_any: dict[str, int] = {}
    for pair in logical_pairs:
        if is_package_title(pair["title"]):
            key = strip_pmd_marker(pair["title"])
            package_history_any[key] = package_history_any.get(key, 0) + 1

    package_occurrences: list[dict[str, Any]] = []
    current_package: Optional[dict[str, Any]] = None
    current_action: Optional[dict[str, Any]] = None

    for pair in active_pairs:
        title = pair["title"]
        zyklus = pair["zyklentext"]

        if is_package_title(title):
            current_package = {
                "package": strip_pmd_marker(title),
                "package_raw": title,
                "is_pmd": "(pmd)" in title.lower(),
                "origin": "PMD" if "(pmd)" in title.lower() else "manuell",
                "title_row": pair["title_row"],
                "status_row": pair["status_row"],
                "package_details": [],
                "actions": [],
            }
            package_occurrences.append(current_package)
            current_action = None
            continue

        if current_package is None:
            current_package = {
                "package": "(ohne Paket)",
                "package_raw": "(ohne Paket)",
                "is_pmd": False,
                "origin": "manuell",
                "title_row": pair["title_row"],
                "status_row": pair["status_row"],
                "package_details": [],
                "actions": [],
            }
            package_occurrences.append(current_package)
            current_action = None

        if zyklus or is_action_title(title):
            current_action = {
                "title": title,
                "title_row": pair["title_row"],
                "status_row": pair["status_row"],
                "zyklentext": zyklus,
                "details": [],
                "title_key": normalize_title_key(title),
            }
            current_package["actions"].append(current_action)
        else:
            detail_obj = {
                "title": title,
                "title_row": pair["title_row"],
                "status_row": pair["status_row"],
            }

            if current_action is not None:
                current_action["details"].append(detail_obj)
            else:
                current_package["package_details"].append(detail_obj)

    latest_by_package: dict[str, dict[str, Any]] = {}
    for occurrence in package_occurrences:
        latest_by_package[occurrence["package"]] = occurrence

    kept_packages = list(latest_by_package.values())
    kept_packages.sort(key=lambda pkg: package_sort_key(pkg["package"]))

    entries: list[dict[str, Any]] = []
    id_counter = {"value": 0}

    for package_obj in kept_packages:
        package_name = package_obj["package"]
        package_raw = package_obj["package_raw"]
        package_id = _next_id(id_counter)

        # Zähle gleiche Leistungstitel pro Paket, damit Instanzen markiert werden können.
        counts_by_title: dict[str, int] = {}
        for action in package_obj["actions"]:
            counts_by_title[action["title_key"]] = counts_by_title.get(action["title_key"], 0) + 1

        seen_by_title: dict[str, int] = {}

        entries.append(
            {
                "entry_id": package_id,
                "parent_entry_id": None,
                "row_type": "paket",
                "title": package_raw,
                "display_title": package_raw,
                "package": package_name,
                "package_raw": package_raw,
                "parent_leistung": None,
                "parent_display": None,
                "zyklentext": "",
                "zyklus_info": parse_zyklentext(""),
                "package_is_pmd": package_obj["is_pmd"],
                "entry_origin": package_obj["origin"],
                "status": "Aktiv",
                "title_row": package_obj["title_row"],
                "status_row": package_obj["status_row"],
                "is_planned_core": True,
                "package_seen_anywhere_in_stay": package_history_any.get(package_name, 0) > 0,
                "package_occurrences_in_stay": package_history_any.get(package_name, 0),
                "instance_index": 1,
                "instance_total_same_title": 1,
            }
        )

        for package_detail in package_obj["package_details"]:
            entries.append(
                {
                    "entry_id": _next_id(id_counter),
                    "parent_entry_id": package_id,
                    "row_type": "detail",
                    "title": package_detail["title"],
                    "display_title": package_detail["title"],
                    "package": package_name,
                    "package_raw": package_raw,
                    "parent_leistung": None,
                    "parent_display": package_raw,
                    "zyklentext": "",
                    "zyklus_info": parse_zyklentext(""),
                    "package_is_pmd": package_obj["is_pmd"],
                    "entry_origin": package_obj["origin"],
                    "status": "Aktiv",
                    "title_row": package_detail["title_row"],
                    "status_row": package_detail["status_row"],
                    "is_planned_core": False,
                    "package_seen_anywhere_in_stay": package_history_any.get(package_name, 0) > 0,
                    "package_occurrences_in_stay": package_history_any.get(package_name, 0),
                    "instance_index": 1,
                    "instance_total_same_title": 1,
                }
            )

        for action in package_obj["actions"]:
            seen_by_title[action["title_key"]] = seen_by_title.get(action["title_key"], 0) + 1
            idx = seen_by_title[action["title_key"]]
            total = counts_by_title[action["title_key"]]

            display_title = action["title"]
            if total > 1:
                display_title = f"{action['title']} [{idx}/{total}]"

            action_id = _next_id(id_counter)

            entries.append(
                {
                    "entry_id": action_id,
                    "parent_entry_id": package_id,
                    "row_type": "leistung",
                    "title": action["title"],
                    "display_title": display_title,
                    "package": package_name,
                    "package_raw": package_raw,
                    "parent_leistung": None,
                    "parent_display": package_raw,
                    "zyklentext": action["zyklentext"],
                    "zyklus_info": parse_zyklentext(action["zyklentext"]),
                    "package_is_pmd": package_obj["is_pmd"],
                    "entry_origin": package_obj["origin"],
                    "status": "Aktiv",
                    "title_row": action["title_row"],
                    "status_row": action["status_row"],
                    "is_planned_core": True,
                    "package_seen_anywhere_in_stay": package_history_any.get(package_name, 0) > 0,
                    "package_occurrences_in_stay": package_history_any.get(package_name, 0),
                    "instance_index": idx,
                    "instance_total_same_title": total,
                }
            )

            for detail in action["details"]:
                entries.append(
                    {
                        "entry_id": _next_id(id_counter),
                        "parent_entry_id": action_id,
                        "row_type": "detail",
                        "title": detail["title"],
                        "display_title": detail["title"],
                        "package": package_name,
                        "package_raw": package_raw,
                        "parent_leistung": action["title"],
                        "parent_display": display_title,
                        "zyklentext": "",
                        "zyklus_info": parse_zyklentext(""),
                        "package_is_pmd": package_obj["is_pmd"],
                        "entry_origin": package_obj["origin"],
                        "status": "Aktiv",
                        "title_row": detail["title_row"],
                        "status_row": detail["status_row"],
                        "is_planned_core": False,
                        "package_seen_anywhere_in_stay": package_history_any.get(package_name, 0) > 0,
                        "package_occurrences_in_stay": package_history_any.get(package_name, 0),
                        "instance_index": idx,
                        "instance_total_same_title": total,
                    }
                )

    stats: dict[str, Any] = {
        "version": "v08.1",
        "sheet_name": worksheet.title,
        "header_row_detected": header_row,
        "logical_pairs_total": len(logical_pairs),
        "logical_pairs_aktiv": len(active_pairs),
        "logical_pairs_abgesetzt": len(abgesetzt_pairs),
        "package_occurrences_aktiv": len(package_occurrences),
        "unique_packages_kept_latest": len(kept_packages),
        "entries_output": len(entries),
        "package_history_any": package_history_any,
        "format_note": "Zweizeilen-Format; Details hängen in v08 an eindeutigen Leistungsinstanzen.",
    }

    return entries, stats


def export_to_json(data: List[Dict[str, Any]], output_path: str | Path) -> None:
    Path(output_path).write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


if __name__ == "__main__":
    print("Dies ist der Parser v08.1 von PRISMA.")
    print("Bitte starte die Streamlit-Oberfläche: prisma_app_v09.py")
