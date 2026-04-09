"""
PRISMA - XLSX Parser
====================

Stabile PRISMA-Basis für den XLSX-Import aus dem KIS.

Ziele:
- nur XLSX als Eingabequelle
- Status "Abgesetzt" konsequent ausblenden
- "Angelegt von" ignorieren
- Zyklentext erhalten und vorinterpretieren
- PMD auf Paketebene erkennen
- Paket / Leistung / Detail sauber trennen

Hinweis:
Diese Version ist absichtlich konservativ.
Sie bevorzugt Stabilität vor zu viel "Magie".
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ============================================================
# KONFIGURATION
# ============================================================

HEADER_ROW = 5
FIRST_DATA_ROW = 6

COL_TITLE = 1
COL_STATUS = 2
COL_EVAL = 3
COL_ZYKLUS = 4
COL_ANLEGEDAT = 5
COL_ERFZEIT = 6
COL_ANGELEGT_VON = 7


# ============================================================
# HILFSFUNKTIONEN
# ============================================================

def clean_text(value: Any) -> str:
    """Normalisiert Zellinhalte zu sauberem Text."""
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def strip_pmd_marker(title: str) -> str:
    """Entfernt die Anzeige-Markierung '(PMD)' aus einem Paketnamen."""
    return re.sub(r"\s*\(PMD\)\s*$", "", title).strip()


def get_indent_level(ws, row_idx: int) -> int:
    """Liest die visuelle Einrückung aus Spalte A."""
    indent = ws.cell(row_idx, COL_TITLE).alignment.indent
    if indent is None:
        return 0
    return int(indent)


def classify_row_type(indent_level: int) -> str:
    """
    Ordnet eine Zeile anhand der Einrückung einem Typ zu.

    Bekannte Muster in der Testdatei:
    - 2 = Paket
    - 4 = Leistung
    - 6 = Detail/Parameter
    """
    if indent_level <= 2:
        return "paket"
    if indent_level == 4:
        return "leistung"
    if indent_level >= 6:
        return "detail"
    return "unbekannt"


def parse_zyklentext(zyklentext: str) -> Dict[str, Any]:
    """
    Interpretiert häufige Zyklustexte in eine einfache Struktur.
    """
    text = clean_text(zyklentext)

    result: Dict[str, Any] = {
        "raw": text,
        "kind": None,
        "frequency_per_day": None,
        "times": [],
        "interval_hours": None,
        "notes": None,
    }

    if not text:
        return result

    lower = text.lower()

    if "bei bedarf" in lower:
        result["kind"] = "bei_bedarf"
        return result

    if lower.startswith("einmalig"):
        result["kind"] = "einmalig"
        result["notes"] = text
        return result

    match = re.search(r"(\d+)x\s*(?:tgl\.|pro tag)", lower)
    if match:
        result["kind"] = "mehrfach_pro_tag"
        result["frequency_per_day"] = int(match.group(1))
        result["times"] = re.findall(r"\b\d{1,2}:\d{2}\b", text)
        if not result["times"]:
            result["notes"] = text
        return result

    match = re.search(r"alle\s+(\d+)\s+stunden", lower)
    if match:
        result["kind"] = "stundenintervall"
        result["interval_hours"] = int(match.group(1))
        result["notes"] = text
        return result

    if lower.startswith("zyklus für die eqs-erfassung"):
        result["kind"] = "sonderzyklus"
        result["notes"] = text
        return result

    result["kind"] = "frei_text"
    result["notes"] = text
    return result


# ============================================================
# PARSER
# ============================================================

def parse_pflegeplan_xlsx(xlsx_path: str | Path) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """
    Liest den Pflegeplan aus einer Excel-Datei und gibt zurück:
    1. aktive Einträge
    2. Parser-Statistik

    Regel:
    - Nur Status 'Aktiv' wird verarbeitet
    - Status 'Abgesetzt' wird komplett ignoriert
    """
    workbook = load_workbook(xlsx_path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]

    parsed_rows: List[Dict[str, Any]] = []

    current_package: Optional[Dict[str, Any]] = None
    current_leistung: Optional[Dict[str, Any]] = None

    stats = {
        "rows_total": 0,
        "rows_empty": 0,
        "rows_ignored_no_title": 0,
        "rows_abgesetzt": 0,
        "rows_aktiv": 0,
        "rows_other_status": 0,
        "entries_output": 0,
    }

    for row_idx in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        raw_title = worksheet.cell(row_idx, COL_TITLE).value
        raw_status = worksheet.cell(row_idx, COL_STATUS).value

        title = clean_text(raw_title)
        status = clean_text(raw_status)
        status_lower = status.lower()

        if not title and not status:
            stats["rows_empty"] += 1
            continue

        stats["rows_total"] += 1

        if not title:
            stats["rows_ignored_no_title"] += 1
            continue

        # ----------------------------------------------------
        # STATUSFILTER
        # ----------------------------------------------------
        if "abgesetzt" in status_lower:
            stats["rows_abgesetzt"] += 1
            continue

        if "aktiv" not in status_lower:
            stats["rows_other_status"] += 1
            continue

        stats["rows_aktiv"] += 1

        # ----------------------------------------------------
        # ZEILENKLASSIFIKATION
        # ----------------------------------------------------
        indent_level = get_indent_level(worksheet, row_idx)
        row_type = classify_row_type(indent_level)

        zyklus_raw = clean_text(worksheet.cell(row_idx, COL_ZYKLUS).value)
        zyklus_info = parse_zyklentext(zyklus_raw)

        # ----------------------------------------------------
        # KONTEXT AKTUALISIEREN
        # ----------------------------------------------------
        if row_type == "paket":
            is_pmd = "(PMD)" in title
            current_package = {
                "title_raw": title,
                "title_clean": strip_pmd_marker(title),
                "is_pmd": is_pmd,
                "row": row_idx,
            }
            current_leistung = None

        elif row_type == "leistung":
            current_leistung = {
                "title": title,
                "row": row_idx,
            }

        entry = {
            "row_index": row_idx,
            "title": title,
            "status": status,
            "row_type": row_type,
            "indent_level": indent_level,
            "package": current_package["title_clean"] if current_package else None,
            "package_raw": current_package["title_raw"] if current_package else None,
            "package_is_pmd": current_package["is_pmd"] if current_package else False,
            "entry_origin": "PMD" if (current_package and current_package["is_pmd"]) else "manuell",
            "parent_leistung": current_leistung["title"] if (row_type == "detail" and current_leistung) else None,
            "zyklentext": zyklus_raw,
            "zyklus_info": zyklus_info,
            "is_planned_core": row_type in {"paket", "leistung"},
        }

        parsed_rows.append(entry)

    stats["entries_output"] = len(parsed_rows)
    return parsed_rows, stats


def export_to_json(data: List[Dict[str, Any]], output_path: str | Path) -> None:
    """Speichert die Parser-Ausgabe als JSON-Datei."""
    Path(output_path).write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


if __name__ == "__main__":
    source_file = Path("TEst Pflegeplan.xlsx")
    output_file = Path("pflegeplan_parsed_preview.json")

    entries, stats = parse_pflegeplan_xlsx(source_file)
    export_to_json(entries, output_file)

    print(f"Fertig. {len(entries)} aktive Einträge exportiert nach: {output_file}")
    print("Parser-Statistik:")
    for key, value in stats.items():
        print(f"  {key}: {value}")
